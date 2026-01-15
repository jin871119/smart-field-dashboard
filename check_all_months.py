import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['단체']

print("=== 더현대서울의 25년 모든 월 데이터 찾기 ===")

# 25년 1~11월
months_2025 = [f'2025{str(m).zfill(2)}' for m in range(1, 12)]

total_sales = 0
found_months = []

for row_idx in range(2, sheet.max_row + 1):
    store_name = sheet.cell(row=row_idx, column=2).value
    if store_name and '더현대서울' in str(store_name):
        date_value = sheet.cell(row=row_idx, column=3).value
        small_group_sales = sheet.cell(row=row_idx, column=20).value or 0
        
        # 날짜를 문자열로 변환
        date_str = str(date_value) if date_value else ''
        
        # 25년 1~11월 확인
        if date_str in months_2025:
            sales_amount = float(small_group_sales) if isinstance(small_group_sales, (int, float)) else 0
            total_sales += sales_amount
            found_months.append({
                'date': date_str,
                'sales': sales_amount
            })
            print(f"  {date_str}: {sales_amount:,.0f}원")

print(f"\n발견된 월: {len(found_months)}개")
print(f"합계: {total_sales:,.0f}원")
print(f"합계 (만원): {total_sales / 10000:,.0f}만원")

# 기대값과 비교
expected = 2904269370
print(f"\n기대값: {expected:,.0f}원")
print(f"계산값: {total_sales:,.0f}원")
print(f"차이: {abs(total_sales - expected):,.0f}원")


