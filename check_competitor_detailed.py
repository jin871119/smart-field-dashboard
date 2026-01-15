# -*- coding: utf-8 -*-
import openpyxl

excel_file = 'backdata.xlsx'

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['경쟁사']
    
    print("Row 2 브랜드 확인 (Column 13부터):")
    brands_found = []
    for col_idx in range(13, min(50, sheet.max_column + 1)):
        cell = sheet.cell(row=2, column=col_idx)
        if cell.value:
            brands_found.append((col_idx, cell.value))
            print(f"  Col {col_idx}: {cell.value}")
    
    print(f"\n총 {len(brands_found)}개 브랜드 발견")
    
    print("\nRow 4 데이터 확인 (K열 백화점 이름):")
    for row_idx in range(4, min(15, sheet.max_row + 1)):
        k_col = sheet.cell(row=row_idx, column=11).value
        print(f"  Row {row_idx}: {k_col}")
        
        # 브랜드 데이터도 확인 (첫 번째 브랜드 컬럼)
        if brands_found:
            first_brand_col = brands_found[0][0]
            brand_value = sheet.cell(row=row_idx, column=first_brand_col).value
            print(f"    -> Col {first_brand_col} ({brands_found[0][1]}): {brand_value}")
    
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
