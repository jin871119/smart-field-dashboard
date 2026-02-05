# -*- coding: utf-8 -*-
import openpyxl
import json
import os
from datetime import datetime

EXCEL_FILE = 'backdata.xlsx'
DATA_DIR = 'public/data'
if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR)

def format_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return value

def normalize_store_name(name):
    if not name:
        return name
    name_str = str(name).strip()
    
    # Specific normalization for Ulsan stores
    # '현대울산' is likely the main branch, '현대울산동구' is the Dong-gu branch
    if name_str in ['현대울산(동)', '현대울산(동구)', '현대 울산 동구']:
        return '현대울산동구'
    if name_str == '현대울산':
        return '현대울산'
    
    return name_str

def process_performance_sheet(workbook, sheet_name, output_filename):
    if sheet_name not in workbook.sheetnames:
        print(f"Skipping: '{sheet_name}' sheet not found.")
        return False
    
    print(f"Processing performance sheet (with aggregation): {sheet_name}...")
    sheet = workbook[sheet_name]
    
    # Read headers
    headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
    
    # Index of key columns
    try:
        period_idx = headers.index('판매시점')
        name_idx = headers.index('매장명')
        sales_idx = headers.index('판매액')
    except ValueError as e:
        print(f"Error: Missing required header in {sheet_name}: {e}")
        return False

    # Aggregate data by (period, normalized_name)
    aggregated_data = {}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[period_idx] is None:
            continue
            
        period = str(row[period_idx]).strip()
        raw_name = row[name_idx]
        normalized_name = normalize_store_name(raw_name)
        
        try:
            sales = float(row[sales_idx]) if row[sales_idx] is not None else 0
        except:
            sales = 0
            
        key = (period, normalized_name)
        if key not in aggregated_data:
            aggregated_data[key] = 0
        aggregated_data[key] += sales
        
    # Convert back to list format for JSON
    data_list = []
    for (period, name), sales in aggregated_data.items():
        data_list.append({
            '판매시점': period,
            '매장명': name,
            '판매액': sales
        })
    
    # Sort for predictability
    data_list.sort(key=lambda x: (x['판매시점'], x['매장명']))
    
    output_path = os.path.join(DATA_DIR, output_filename)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({
            'headers': ['판매시점', '매장명', '판매액'],
            'data': data_list,
            'total_rows': len(data_list)
        }, f, ensure_ascii=False, indent=2)
    
    print(f"Saved {len(data_list)} aggregated rows to {output_filename}")
    return True

def process_generic_sheet(workbook, sheet_name, output_filename):
    if sheet_name not in workbook.sheetnames:
        print(f"Skipping: '{sheet_name}' sheet not found.")
        return False
    
    print(f"Processing generic sheet: {sheet_name}...")
    sheet = workbook[sheet_name]
    
    # Read headers
    headers = [cell.value for cell in sheet[1]]
    data = []
    
    # Read data starting from row 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        has_data = False
        for header, value in zip(headers, row):
            val = format_date(value)
            row_data[header] = val
            if val is not None and val != '':
                has_data = True
        
        if has_data:
            data.append(row_data)
    
    output_path = os.path.join(DATA_DIR, output_filename)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump({
            'headers': headers,
            'data': data,
            'total_rows': len(data)
        }, f, ensure_ascii=False, indent=2)
    
    print(f"Saved {len(data)} rows to {output_filename}")
    return True

def process_group_sales(workbook, sheet_name, output_filename):
    if sheet_name not in workbook.sheetnames:
        print(f"Skipping: '{sheet_name}' sheet not found.")
        return False
    
    print(f"Processing group sales (special): {sheet_name}...")
    sheet = workbook[sheet_name]
    store_map = {}
    months_2025 = [f'2025{str(m).zfill(2)}' for m in range(1, 12)]
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Column 2: Store Name, Column 3: Date, Column 20: Sales (T column)
        store_name = row[1]
        date_val = row[2]
        sales_val = row[19] # 0-indexed, so T is 19
        
        if not store_name:
            continue
            
        store_name = str(store_name).strip()
        date_str = str(date_val) if date_val else ''
        
        if date_str not in months_2025:
            continue
            
        try:
            sales = float(sales_val) if sales_val is not None else 0
        except:
            sales = 0
            
        if store_name not in store_map:
            store_map[store_name] = {'매장명': store_name, '소량단체판매액': 0}
        
        store_map[store_name]['소량단체판매액'] += sales
        
    result = {
        'stores': list(store_map.values()),
        'total_stores': len(store_map)
    }
    
    output_path = os.path.join(DATA_DIR, output_filename)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"Saved {len(store_map)} stores to {output_filename}")
    return True

def process_competitor(workbook, sheet_name, output_filename):
    if sheet_name not in workbook.sheetnames:
        print(f"Skipping: '{sheet_name}' sheet not found.")
        return False
        
    print(f"Processing competitor (special): {sheet_name}...")
    sheet = workbook[sheet_name]
    # Row 1 and 2 are headers
    row1 = [cell.value for cell in sheet[1]]
    row2 = [cell.value for cell in sheet[2]]
    
    # Identifiy品牌的范围
    avg_col_indices = []
    for col_idx, (h1, h2) in enumerate(zip(row1, row2)):
        h1_str = str(h1) if h1 else ""
        h2_str = str(h2) if h2 else ""
        if '월평균' in h1_str or '월평균' in h2_str:
            avg_col_indices.append(col_idx)

    if not avg_col_indices:
        print("Error: Could not find '월평균' column in competitor sheet.")
        return False
    
    # Re-writing the brand identification logic more robustly
    brands = []
    
    # The range is from first '월평균' to the last one, or until next category
    # Actually, let's just find ALL brands mentioned in row 2 that are under the '월평균' columns
    # In this file, brands start from Col 13.
    # Let's hardcode the range if necessary, or detect it:
    start_col = avg_col_indices[0]
    # Find brands in row 2 for these columns
    for col_idx in range(start_col, start_col + 15): # Assume max 15 brands in first avg section
        brand_name = str(row2[col_idx]).strip() if row2[col_idx] else None
        if brand_name and brand_name != 'None' and brand_name != 'MLB' and col_idx not in [b['col_idx'] for b in brands]:
             # If MLB was already added or not
             pass
        
    # Re-writing the brand identification logic more robustly
    brands = []
    # MLB is always first at Col 13 (index 12)
    current_brand_row = row2 # Usually brands are in row 2
    
    # Let's find columns where row 1 has '월평균' and row 2 has brand name
    # OR row 1 is empty (merged) and row 2 has brand name
    is_in_avg_section = False
    for col_idx in range(len(row1)):
        h1 = str(row1[col_idx]) if row1[col_idx] else ""
        h2 = str(row2[col_idx]).strip() if row2[col_idx] and row2[col_idx] != 'None' else ""
        
        if '월평균' in h1:
            is_in_avg_section = True
        elif h1 != "" and '월평균' not in h1: # Next section started
             # But wait, there might be multiple 월평균 sections (like Monthly MS)
             # Let's only take the first section '월평균(1~12월)'
             if is_in_avg_section:
                 break
        
        if is_in_avg_section and h2:
            brands.append({'col_idx': col_idx, 'name': h2})

    # Also look for a total sales column (usually '매출액' or similar)
    total_sales_col = -1
    for col_idx, h1 in enumerate(row1):
        if h1 and '매출액' in str(h1) and '합계' in str(h1):
            total_sales_col = col_idx
            break
    
    stores_data = []
    # [간소화된 시트] A열(백화점), B~=월평균 브랜드 데이터. Row 3부터 데이터
    STORE_COL_IDX = 0  # A열 = 백화점
    for row in sheet.iter_rows(min_row=3, values_only=True):
        store_name = row[STORE_COL_IDX] if len(row) > STORE_COL_IDX else None
        
        # Skip empty rows or rows that are just headers
        if not store_name or str(store_name).strip() == '' or str(store_name).strip() == '백화점':
            continue
        store_name = str(store_name).strip()
        if store_name.isdigit():  # 숫자만 있는 행(인덱스 등) 제외
            continue
            
        brand_data = {}
        total_brand_sales = 0
        for brand in brands:
            val = row[brand['col_idx']]
            try:
                val_float = float(val) if val is not None else 0
                brand_data[brand['name']] = val_float
                total_brand_sales += val_float
            except:
                brand_data[brand['name']] = 0
                
        stores_data.append({
            '백화점': store_name,
            '브랜드별_월평균': brand_data,
            '총매출': total_brand_sales # Use the sum of brands as the store total
        })
        
    result = {
        'brands': [b['name'] for b in brands],
        'stores': stores_data,
        'total_stores': len(stores_data),
        'has_total_sales': total_sales_col >= 0
    }
    
    output_path = os.path.join(DATA_DIR, output_filename)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
        
    print(f"Saved {len(stores_data)} stores to {output_filename}")
    return True

if __name__ == '__main__':
    print("=" * 50)
    print("Starting Optimized Dashboard Data Update")
    print("=" * 50)
    
    start_time = datetime.now()
    
    try:
        print(f"Loading {EXCEL_FILE} (this may take a minute color 64MB)...")
        # data_only=True to get calculated values
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        print(f"File loaded in {datetime.now() - start_time}")
        
        # Standard sheets
        process_generic_sheet(wb, '매장', 'store_data.json')
        process_generic_sheet(wb, '아이템시즌별판매', 'item_season_data.json')
        
        # Optimize '매장별스타일판매' to reduce file size (it was 180MB+)
        print("Processing optimized sheet: 매장별스타일판매...")
        style_sheet = wb['매장별스타일판매']
        style_headers = [cell.value for cell in style_sheet[1]]
        
        # Define mapping from possible Excel headers to JSON keys
        header_mapping = {
            '매장명': '매장명',
            '품번': '품번',
            '제품명': '제품명',
            '판매액': '판매액합계',
            '판매액합계': '판매액합계',
            '판매수량': '판매수량합계',
            '판매수량합계': '판매수량합계',
            '일자': '일자',
            '시즌': '시즌'
        }
        
        # Build col_indices based on available headers
        col_indices = {}
        for excel_h, json_h in header_mapping.items():
            if excel_h in style_headers:
                col_indices[json_h] = style_headers.index(excel_h)
        
        date_idx = style_headers.index('일자') if '일자' in style_headers else -1
        print(f"DEBUG: col_indices={col_indices}, date_idx={date_idx}")
        
        style_data = []
        for i, row in enumerate(style_sheet.iter_rows(min_row=2, values_only=True)):
            # If '일자' exists, filter for 2026. Otherwise include all (per user request)
            include_row = True
            if date_idx >= 0:
                val = row[date_idx]
                try:
                    v_str = str(val)
                    if '2026' not in v_str:
                        include_row = False
                except:
                    pass
            
            if not include_row:
                continue
                
            row_data = {}
            for json_h, idx in col_indices.items():
                val = row[idx]
                if json_h == '매장명':
                    val = normalize_store_name(val)
                row_data[json_h] = format_date(val)
            style_data.append(row_data)
            
        with open(os.path.join(DATA_DIR, 'store_style_sales_data.json'), 'w', encoding='utf-8') as f:
            json.dump({
                'headers': list(col_indices.keys()),
                'data': style_data,
                'total_rows': len(style_data)
            }, f, ensure_ascii=False, indent=2)
        print(f"Saved {len(style_data)} optimized rows to store_style_sales_data.json")

        process_generic_sheet(wb, '매장별재고', 'store_inventory_data.json')
        process_performance_sheet(wb, '실적', 'performance_data.json')
        # process_generic_sheet(wb, '주간회의', 'weekly_meeting_data.json') # Removed as it's missing in current file
        
        # Specialized sheets
        process_group_sales(wb, '단체', 'group_sales_data.json')
        process_competitor(wb, '경쟁사', 'competitor_data_v2.json')
        
        print("\n" + "=" * 50)
        print(f"All data updated successfully in {datetime.now() - start_time}")
        print("Output directory: public/data/")
        print("=" * 50)
        
    except Exception as e:
        print(f"CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()
