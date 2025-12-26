import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx')
sheet = wb['경쟁사']

print("=== Row 1 (헤더) ===")
row1 = [cell.value for cell in sheet[1][:20]]
for i, val in enumerate(row1):
    print(f"Column {i+1} (chr {chr(65+i)}): {val}")

print("\n=== Row 2 (브랜드) ===")
row2 = [cell.value for cell in sheet[2][:20]]
for i, val in enumerate(row2):
    print(f"Column {i+1} (chr {chr(65+i)}): {val}")

print("\n=== Row 3 (데이터 시작) ===")
row3 = [cell.value for cell in sheet[3][:20]]
for i, val in enumerate(row3):
    print(f"Column {i+1} (chr {chr(65+i)}): {val}")

print("\n=== Column K (11번째 컬럼) 값들 ===")
for i in range(1, 10):
    val = sheet.cell(row=i, column=11).value
    print(f"Row {i}: {val}")

print("\n=== 월평균 컬럼 찾기 ===")
# 월평균(1~11월) 컬럼 찾기
for col_idx, cell in enumerate(sheet[1], 1):
    if cell.value and '월평균' in str(cell.value):
        print(f"월평균 컬럼: Column {col_idx} (chr {chr(64+col_idx)}), Row 1 값: {cell.value}")
        # Row 2의 해당 컬럼 값 (브랜드)
        brand_cell = sheet.cell(row=2, column=col_idx)
        print(f"  Row 2 (브랜드): {brand_cell.value}")
        # 몇 개 데이터 확인
        print(f"  처음 5개 데이터:")
        for row_idx in range(3, 8):
            data_cell = sheet.cell(row=row_idx, column=col_idx)
            store_name = sheet.cell(row=row_idx, column=3).value  # 백화점 이름은 C열 (3번째)
            print(f"    Row {row_idx} ({store_name}): {data_cell.value}")


