import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx')
sheet = wb['경쟁사']

print("=== Row 2 전체 확인 (컬럼 11번부터 30번까지) ===")
for col_idx in range(11, 31):
    cell = sheet.cell(row=2, column=col_idx)
    col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)
    print(f"Column {col_idx} ({col_letter}): {cell.value} (type: {type(cell.value)})")

print("\n=== Row 1에서 월평균 컬럼 찾기 ===")
for col_idx in range(1, sheet.max_column + 1):
    cell = sheet.cell(row=1, column=col_idx)
    if cell.value and '월평균' in str(cell.value):
        print(f"Column {col_idx}: {cell.value}")
        # 해당 컬럼의 Row 2 값 확인
        row2_cell = sheet.cell(row=2, column=col_idx)
        print(f"  Row 2 값: {row2_cell.value}")
        # Row 3 값도 확인
        row3_cell = sheet.cell(row=3, column=col_idx)
        store_name = sheet.cell(row=3, column=11).value
        print(f"  Row 3 ({store_name}): {row3_cell.value}")



