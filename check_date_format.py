import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['단체']

print("=== 더현대서울의 모든 행 확인 ===")

for row_idx in range(2, min(500, sheet.max_row + 1)):
    store_name = sheet.cell(row=row_idx, column=2).value
    if store_name and '더현대서울' in str(store_name):
        date_value = sheet.cell(row=row_idx, column=3).value
        small_group_sales = sheet.cell(row=row_idx, column=20).value
        
        print(f"Row {row_idx}: 날짜={date_value} (type: {type(date_value)}), 소량단체판매액={small_group_sales}")
        
        # 날짜 값이 문자열일 수도 있음
        if date_value:
            date_str = str(date_value)
            print(f"  날짜 문자열: {date_str}")

