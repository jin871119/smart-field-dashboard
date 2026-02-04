# -*- coding: utf-8 -*-
import openpyxl

EXCEL_FILE = 'backdata.xlsx'

try:
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    print("Sheets:", wb.sheetnames)
    
    # Check '경쟁사' sheet
    sheet_name = '경쟁사'
    if sheet_name not in wb.sheetnames:
        # Try with quotes if they are literally in the name
        quoted_name = f"'{sheet_name}'"
        if quoted_name in wb.sheetnames:
            sheet_name = quoted_name
    
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nStructure of '{sheet_name}':")
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True)):
            print(f"Row {i+1}:", row[:15])
    else:
        print(f"\nSheet '{sheet_name}' not found!")

except Exception as e:
    print("Error:", e)
