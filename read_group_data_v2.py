import openpyxl
import json

def read_group_data(file_path, output_json_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        if '단체' not in workbook.sheetnames:
            print("'단체' 시트를 찾을 수 없습니다.")
            return
        
        sheet = workbook['단체']
        
        # 매장별 소량단체판매액 데이터 수집 (25년 1~11월 합계)
        # 각 행이 월별 데이터이므로 매장별로 그룹화하고 합산
        store_map = {}
        
        # 25년 1~11월 날짜 목록
        months_2025 = [f'2025{str(m).zfill(2)}' for m in range(1, 12)]
        
        for row_idx in range(2, sheet.max_row + 1):
            # 매장명 (Column 2)
            store_name = sheet.cell(row=row_idx, column=2).value
            
            # 매장명이 없으면 건너뛰기
            if not store_name:
                continue
            
            store_name = str(store_name).strip()
            
            # 날짜 (Column 3) - 문자열 형식 (예: "202501")
            date_value = sheet.cell(row=row_idx, column=3).value
            date_str = str(date_value) if date_value else ''
            
            # 25년 1~11월 데이터만 처리
            if date_str not in months_2025:
                continue
            
            # 소량단체판매액 (T열 = Column 20)
            small_group_sales = sheet.cell(row=row_idx, column=20).value
            
            # 숫자 값 처리
            if small_group_sales is None:
                small_group_sales = 0
            elif isinstance(small_group_sales, (int, float)):
                small_group_sales = float(small_group_sales)
            else:
                try:
                    small_group_sales = float(small_group_sales)
                except:
                    small_group_sales = 0
            
            # 매장별로 합산
            if store_name not in store_map:
                store_map[store_name] = {
                    '매장명': store_name,
                    '소량단체판매액': 0
                }
            
            store_map[store_name]['소량단체판매액'] += small_group_sales
        
        # 리스트로 변환
        stores_data = list(store_map.values())
        
        # 결과 저장
        result = {
            'stores': stores_data,
            'total_stores': len(stores_data)
        }
        
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"총 {len(stores_data)}개 매장의 데이터를 {output_json_path}에 저장했습니다.")
        print(f"\n처음 5개 매장 데이터:")
        for store in stores_data[:5]:
            print(f"  {store['매장명']}: {store['소량단체판매액']:,.0f}원 ({store['소량단체판매액']/10000:,.0f}만원)")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()

excel_file = 'backdata.xlsx'
output_json_path = 'group_sales_data.json'

read_group_data(excel_file, output_json_path)

