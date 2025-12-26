import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['경쟁사']

# Row 1에서 "월평균"이 있는 컬럼 찾기
print("=== Row 1에서 월평균 컬럼 찾기 ===")
monthly_avg_cols = []
for col_idx in range(1, sheet.max_column + 1):
    header_cell = sheet.cell(row=1, column=col_idx)
    if header_cell.value and '월평균' in str(header_cell.value):
        monthly_avg_cols.append(col_idx)
        print(f"Column {col_idx}: {header_cell.value}")

print(f"\n월평균 컬럼: {monthly_avg_cols}")

# 첫 번째 월평균 컬럼부터 연속으로 브랜드가 있는지 확인
if monthly_avg_cols:
    start_col = monthly_avg_cols[0]
    print(f"\n=== Column {start_col}부터 Row 2의 브랜드 확인 ===")
    brands_found = []
    for col_idx in range(start_col, start_col + 20):  # 최대 20개까지 확인
        brand_cell = sheet.cell(row=2, column=col_idx)
        if brand_cell.value:
            brands_found.append({
                'col': col_idx,
                'name': str(brand_cell.value).strip()
            })
            print(f"Column {col_idx}: {brand_cell.value}")
        elif len(brands_found) > 0:
            # 이미 브랜드를 찾았는데 빈 셀이 나오면 끝
            break

print(f"\n발견된 브랜드 수: {len(brands_found)}")

# Row 3의 데이터 확인 (백화점 이름과 브랜드별 값)
print(f"\n=== Row 3 데이터 확인 ===")
store_name = sheet.cell(row=3, column=11).value  # K열
print(f"백화점 이름 (K열): {store_name}")

if monthly_avg_cols:
    start_col = monthly_avg_cols[0]
    print(f"브랜드별 월평균 값:")
    for col_idx in range(start_col, start_col + 10):
        brand_cell = sheet.cell(row=2, column=col_idx)
        value_cell = sheet.cell(row=3, column=col_idx)
        if brand_cell.value:
            print(f"  Column {col_idx} ({brand_cell.value}): {value_cell.value}")


