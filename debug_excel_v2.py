# -*- coding: utf-8 -*-
import openpyxl

EXCEL_FILE = 'backdata.xlsx'

def safe_str(val):
    if val is None: return "None"
    return str(val)[:30]

try:
    print("Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    print("\nSHEET NAMES:")
    for i, name in enumerate(wb.sheetnames):
        print(f"{i+1}. [{name}] (Length: {len(name)})")
    
    for target in ['경쟁사', '주간회의']:
        print(f"\n--- Checking Sheet: {target} ---")
        best_match = None
        for name in wb.sheetnames:
            if target in name:
                best_match = name
                break
        
        if best_match:
            print(f"Match found: [{best_match}]")
            sheet = wb[best_match]
            for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
                print(f"Row {r_idx+1}:", [safe_str(c) for c in row[:12]])
        else:
            print(f"No match found for '{target}'")

except Exception as e:
    print("Error:", e)
