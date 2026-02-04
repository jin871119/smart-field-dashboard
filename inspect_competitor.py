# -*- coding: utf-8 -*-
import openpyxl

EXCEL_FILE = 'backdata.xlsx'

try:
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    sheet = wb['경쟁사']
    
    # Check Row 1
    row1 = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    print("Row 1 (first 30 cols):", list(row1[:30]))
    
    # Check Row 2
    row2 = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    print("Row 2 (first 30 cols):", list(row2[:30]))
    
    # Check Row 3
    row3 = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
    print("Row 3 (first 30 cols):", list(row3[:30]))
    
    # Find brands
    brands = []
    for i, (h1, h2) in enumerate(zip(row1, row2)):
        if h1 and '월평균' in str(h1):
            brands.append({'col': i+1, 'h1': h1, 'h2': h2})
    
    print("\nPotential Brands:")
    for b in brands:
        print(f"Col {b['col']}: Row1='{b['h1']}', Row2='{b['h2']}'")

except Exception as e:
    print("Error:", e)
