# -*- coding: utf-8 -*-
import openpyxl

excel_file = 'backdata.xlsx'

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['실적']
    
    print("실적 시트 구조 확인:")
    print(f"총 행 수: {sheet.max_row}")
    print(f"총 열 수: {sheet.max_column}")
    
    # Row 1 헤더 확인
    print("\nRow 1 (헤더):")
    headers = []
    for col_idx in range(1, min(15, sheet.max_column + 1)):
        cell = sheet.cell(row=1, column=col_idx)
        headers.append(cell.value)
        if cell.value:
            print(f"  Col {col_idx} ({chr(64+col_idx)}열): {cell.value}")
    
    # Row 2-5 데이터 샘플 확인
    print("\nRow 2-5 데이터 샘플:")
    for row_idx in range(2, min(6, sheet.max_row + 1)):
        print(f"\nRow {row_idx}:")
        b_col = sheet.cell(row=row_idx, column=2).value  # 판매시점
        c_col = sheet.cell(row=row_idx, column=3).value  # 매장명
        j_col = sheet.cell(row=row_idx, column=10).value  # 판매액
        print(f"  B열 (판매시점): {b_col}")
        print(f"  C열 (매장명): {c_col}")
        print(f"  J열 (판매액): {j_col}")
    
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
