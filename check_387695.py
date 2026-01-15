import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['경쟁사']

row = 3  # 롯데잠실

print("=== 387695 근처 값 찾기 ===")
for col_idx in range(1, min(100, sheet.max_column + 1)):
    try:
        value = sheet.cell(row=row, column=col_idx).value
        if value is not None:
            if isinstance(value, (int, float)):
                # 387695와 비슷한 값 찾기
                if abs(float(value) - 387695) < 1000:
                    header1 = sheet.cell(row=1, column=col_idx).value
                    header2 = sheet.cell(row=2, column=col_idx).value
                    print(f"Column {col_idx}: {value} (Row 1: {header1}, Row 2: {header2})")
    except:
        pass

# 합계 컬럼 확인
print(f"\n=== 합계 컬럼 확인 ===")
for col_idx in range(1, min(100, sheet.max_column + 1)):
    header1 = sheet.cell(row=1, column=col_idx).value
    if header1 and '합계' in str(header1):
        value = sheet.cell(row=row, column=col_idx).value
        header2 = sheet.cell(row=2, column=col_idx).value
        print(f"Column {col_idx} ({header1}): {value} (Row 2: {header2})")



