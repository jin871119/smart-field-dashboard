# -*- coding: utf-8 -*-
import openpyxl

excel_file = 'backdata.xlsx'

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['경쟁사']
    
    print("Row 1-3 전체 확인:")
    for row_idx in range(1, 4):
        print(f"\nRow {row_idx}:")
        for col_idx in range(1, 20):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                print(f"  Col {col_idx}: {cell.value}")
    
    print("\n\nRow 4-8 전체 확인 (데이터 행):")
    for row_idx in range(4, 9):
        print(f"\nRow {row_idx}:")
        row_data = []
        for col_idx in range(1, 20):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                row_data.append(f"Col{col_idx}={cell.value}")
        print("  " + ", ".join(row_data))
    
    # 백화점 이름이 있는 열 찾기 (Row 4 기준)
    print("\n\n백화점 이름이 있을 것으로 추정되는 열 (Row 4-10):")
    candidate_cols = []
    for col_idx in range(1, 15):
        values = []
        for row_idx in range(4, 11):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                values.append(str(cell.value))
        if len(set(values)) > 1:  # 서로 다른 값이 여러 개 있으면
            candidate_cols.append((col_idx, values[:3]))
    
    for col_idx, values in candidate_cols:
        print(f"  Col {col_idx}: {values}")
    
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
