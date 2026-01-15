# -*- coding: utf-8 -*-
import openpyxl
import json
from datetime import datetime

excel_file = 'backdata.xlsx'

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    
    # 모든 시트 확인
    print("모든 시트 확인 중...")
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # 시트 이름이나 헤더에 '주간' 또는 '회의'가 있는지 확인
        headers = [str(cell.value) if cell.value else '' for cell in sheet[1]]
        has_weekly_keyword = any('주간' in h or '회의' in h for h in headers) or '주간' in sheet_name or '회의' in sheet_name
        
        if has_weekly_keyword:
            print(f"\n✓ 발견: '{sheet_name}' 시트")
            print(f"  헤더: {headers[:5]}...")  # 처음 5개만
            
            # 데이터 읽기 및 저장
            data = []
            for row_idx in range(2, sheet.max_row + 1):
                row_data = {}
                has_data = False
                
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                    elif value is None:
                        value = None
                    
                    row_data[header] = value
                    
                    if value is not None and value != '':
                        has_data = True
                
                if has_data:
                    data.append(row_data)
            
            # JSON 저장
            result = {
                'headers': headers,
                'data': data,
                'total_rows': len(data)
            }
            
            output_json_path = 'weekly_meeting_data.json'
            with open(output_json_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            
            print(f"  ✓ {len(data)}개 행을 {output_json_path}에 저장했습니다.")
            break
    else:
        print("\n'주간' 또는 '회의' 키워드를 가진 시트를 찾을 수 없습니다.")
        print("\n사용 가능한 시트:")
        for idx, name in enumerate(workbook.sheetnames, 1):
            print(f"  {idx}. {name}")
        print("\n시트 이름을 정확히 알려주시면 해당 시트를 읽겠습니다.")
        
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()

