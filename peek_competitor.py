# -*- coding: utf-8 -*-
import openpyxl

EXCEL_FILE = 'backdata.xlsx'

try:
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    if '경쟁사' in wb.sheetnames:
        sheet = wb['경쟁사']
        print("Dumping first 5 rows of '경쟁사':")
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True)):
            print(f"Row {i+1}:", list(row))
    else:
        print("'경쟁사' sheet not found")
except Exception as e:
    print("Error:", e)
