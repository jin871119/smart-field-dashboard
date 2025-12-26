import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['경쟁사']

# 롯데잠실 찾기 (K열 = 11번째 컬럼)
print("=== 롯데잠실 데이터 찾기 ===")
lotte_jamsil_row = None
for row_idx in range(3, sheet.max_row + 1):
    store_name = sheet.cell(row=row_idx, column=11).value
    if store_name and '롯데잠실' in str(store_name):
        lotte_jamsil_row = row_idx
        print(f"Row {row_idx}: {store_name}")
        break

if lotte_jamsil_row:
    print(f"\n=== 롯데잠실 Row {lotte_jamsil_row}의 MLB 데이터 확인 ===")
    # L열 (12번째 컬럼)이 MLB
    mlb_value = sheet.cell(row=lotte_jamsil_row, column=12).value
    print(f"L열 (Column 12, MLB): {mlb_value}")
    
    # 주변 컬럼도 확인
    print(f"\n주변 컬럼 확인:")
    for col_idx in range(10, 20):
        header_cell = sheet.cell(row=2, column=col_idx)
        value_cell = sheet.cell(row=lotte_jamsil_row, column=col_idx)
        if header_cell.value:
            print(f"  Column {col_idx} ({header_cell.value}): {value_cell.value}")
    
    # Row 1의 헤더도 확인
    print(f"\nRow 1 헤더 확인:")
    for col_idx in range(10, 20):
        header_cell = sheet.cell(row=1, column=col_idx)
        if header_cell.value:
            print(f"  Column {col_idx}: {header_cell.value}")
    
    # 합계 컬럼 확인
    print(f"\n합계 컬럼 찾기:")
    for col_idx in range(1, sheet.max_column + 1):
        header_cell = sheet.cell(row=1, column=col_idx)
        if header_cell.value and '합계' in str(header_cell.value):
            value_cell = sheet.cell(row=lotte_jamsil_row, column=col_idx)
            print(f"  Column {col_idx} ({header_cell.value}): {value_cell.value}")
else:
    print("롯데잠실을 찾을 수 없습니다.")


