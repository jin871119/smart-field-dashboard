# -*- coding: utf-8 -*-
import openpyxl

excel_file = 'backdata.xlsx'

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['경쟁사']
    
    print(f"총 행 수: {sheet.max_row}")
    print(f"총 열 수: {sheet.max_column}")
    
    # Row 1 확인
    print("\nRow 1 (헤더):")
    for col_idx in range(1, min(20, sheet.max_column + 1)):
        cell = sheet.cell(row=1, column=col_idx)
        print(f"  Col {col_idx}: {cell.value}")
    
    # Row 2 확인 (브랜드)
    print("\nRow 2 (브랜드):")
    for col_idx in range(1, min(30, sheet.max_column + 1)):
        cell = sheet.cell(row=2, column=col_idx)
        if cell.value:
            print(f"  Col {col_idx}: {cell.value}")
    
    # Row 3 확인 (데이터 시작)
    print("\nRow 3 (데이터 시작):")
    for col_idx in range(1, min(15, sheet.max_column + 1)):
        cell = sheet.cell(row=3, column=col_idx)
        if cell.value:
            print(f"  Col {col_idx}: {cell.value}")
    
    # K열 (11번째 컬럼) 확인 - 백화점 이름
    print("\nK열 (백화점 이름) - 처음 10개 행:")
    for row_idx in range(3, min(13, sheet.max_row + 1)):
        cell = sheet.cell(row=row_idx, column=11)
        print(f"  Row {row_idx}: {cell.value}")
    
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
