import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['단체']

print("=== 더현대서울의 25년 1~11월 데이터 찾기 ===")

# 더현대서울의 모든 행 찾기
daehyundai_rows = []
for row_idx in range(2, sheet.max_row + 1):
    store_name = sheet.cell(row=row_idx, column=2).value
    if store_name and '더현대서울' in str(store_name):
        date_value = sheet.cell(row=row_idx, column=3).value
        small_group_sales = sheet.cell(row=row_idx, column=20).value
        
        daehyundai_rows.append({
            'row': row_idx,
            'date': date_value,
            'sales': small_group_sales
        })

print(f"더현대서울 총 {len(daehyundai_rows)}개 행 발견")

# 25년 1~11월 데이터만 필터링
year_2025_data = []
for item in daehyundai_rows:
    date_val = item['date']
    if isinstance(date_val, (int, float)):
        # 202501 ~ 202511 범위 확인
        if 202501 <= date_val <= 202511:
            year_2025_data.append(item)
            print(f"  Row {item['row']}: 날짜={date_val}, 소량단체판매액={item['sales']}")

print(f"\n25년 1~11월 데이터: {len(year_2025_data)}개")

# 합계 계산
total_sales = sum(item['sales'] if item['sales'] else 0 for item in year_2025_data)
print(f"\n합계: {total_sales:,.0f}원")
print(f"합계 (만원): {total_sales / 10000:,.0f}만원")

# 기대값과 비교
expected = 2904269370
print(f"\n기대값: {expected:,.0f}원")
print(f"차이: {abs(total_sales - expected):,.0f}원")


