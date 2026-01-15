# -*- coding: utf-8 -*-
import openpyxl
import json
from datetime import datetime

excel_file = 'backdata.xlsx'
sheet_name = '주간회의'

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    
    print(f"사용 가능한 시트: {workbook.sheetnames}")
    
    if sheet_name not in workbook.sheetnames:
        print(f"\n'{sheet_name}' 시트를 찾을 수 없습니다.")
        print("\n사용 가능한 시트 목록:")
        for idx, name in enumerate(workbook.sheetnames, 1):
            print(f"  {idx}. {name}")
    else:
        sheet = workbook[sheet_name]
        
        # 헤더 읽기 (1행)
        headers = [cell.value for cell in sheet[1]]
        print(f"\n헤더 ({len(headers)}개):")
        for idx, header in enumerate(headers, 1):
            print(f"  {idx}. {header}")
        
        # 데이터 읽기 (2행부터)
        data = []
        
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # 날짜 타입 처리
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                elif value is None:
                    value = None
                
                row_data[header] = value
                
                # 값이 있으면 데이터가 있다고 표시
                if value is not None and value != '':
                    has_data = True
            
            # 데이터가 있는 행만 추가
            if has_data:
                data.append(row_data)
        
        # 결과 저장
        result = {
            'headers': headers,
            'data': data,
            'total_rows': len(data)
        }
        
        output_json_path = 'weekly_meeting_data.json'
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"\n총 {len(data)}개 행의 데이터를 {output_json_path}에 저장했습니다.")
        
        print("\n처음 3개 행 데이터:")
        for i, row in enumerate(data[:3]):
            print(f"\n행 {i+1}:")
            for key, value in list(row.items())[:10]:  # 처음 10개 컬럼만
                print(f"  {key}: {value}")
        
except FileNotFoundError:
    print(f"파일을 찾을 수 없습니다: {excel_file}")
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()

