import openpyxl
import json

def read_competitor_data(file_path, output_json_path):
    try:
        # data_only=True로 실제 계산된 값을 읽기
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['경쟁사']

        # [간소화된 시트 구조]
        # Row 1: A=백화점, B=월평균(1~12월) 등
        # Row 2: B~O = 브랜드명 (MLB, 캉골, 라이프워크, ...)
        # Row 4~: A=백화점명, B~O=브랜드별 월평균 (Row 3은 비어있거나 구분행일 수 있음)

        STORE_NAME_COL = 1  # A열 = 백화점
        DATA_START_ROW = 3  # 데이터 시작 행 (빈 행 건너뛰기 위해 반복에서 체크)

        # Row 2에서 브랜드 이름 수집 (월평균 영역: B열(2)부터 '합계' 전까지)
        brands = []
        for col_idx in range(2, sheet.max_column + 1):
            brand_cell = sheet.cell(row=2, column=col_idx).value
            header_cell = sheet.cell(row=1, column=col_idx).value
            # '합계'는 제외, Row 2에 브랜드명이 있는 열만
            if header_cell and '합계' in str(header_cell):
                break
            if brand_cell and str(brand_cell).strip():
                brand_name = str(brand_cell).strip()
                brands.append({'column': col_idx, 'name': brand_name})

        print(f"발견된 브랜드 ({len(brands)}개): {[b['name'] for b in brands]}")

        # 데이터 읽기 (Row 3부터)
        stores_data = []

        for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
            store_name_val = sheet.cell(row=row_idx, column=STORE_NAME_COL).value
            if not store_name_val or not str(store_name_val).strip():
                continue  # 빈 행은 건너뛰기 (연속 5행 빈 경우 중단)
            store_name = str(store_name_val).strip()
            # 숫자만 있으면 건너뛰기 (헤더/인덱스 행)
            if store_name.isdigit():
                continue

            # 각 브랜드별 월평균 데이터 수집
            brand_data = {}
            for brand in brands:
                cell = sheet.cell(row=row_idx, column=brand['column'])
                value = cell.value
                
                # 숫자 값 처리
                if value is None:
                    brand_data[brand['name']] = 0
                elif isinstance(value, (int, float)):
                    brand_data[brand['name']] = float(value)
                else:
                    # 문자열이면 숫자로 변환 시도
                    try:
                        brand_data[brand['name']] = float(value)
                    except:
                        brand_data[brand['name']] = 0
            
            stores_data.append({
                '백화점': store_name,
                '브랜드별_월평균': brand_data
            })
        
        # 결과 저장
        result = {
            'brands': [b['name'] for b in brands],
            'stores': stores_data,
            'total_stores': len(stores_data)
        }
        
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"\n총 {len(stores_data)}개 점포의 데이터를 {output_json_path}에 저장했습니다.")
        print(f"\n처음 3개 점포 데이터:")
        for store in stores_data[:3]:
            print(f"  {store['백화점']}:")
            for brand, value in store['브랜드별_월평균'].items():
                print(f"    {brand}: {value}")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()

excel_file = 'backdata.xlsx'
output_json_path = 'public/data/competitor_data_v2.json'

read_competitor_data(excel_file, output_json_path)



