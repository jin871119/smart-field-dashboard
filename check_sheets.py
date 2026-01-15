import openpyxl

excel_file = 'backdata.xlsx'
workbook = openpyxl.load_workbook(excel_file)

print("사용 가능한 모든 시트:")
for idx, sheet_name in enumerate(workbook.sheetnames, 1):
    print(f"  {idx}. {repr(sheet_name)}")

# '주간'이 포함된 시트 찾기
weekly_sheets = [name for name in workbook.sheetnames if '주간' in name or 'weekly' in name.lower() or 'meeting' in name.lower()]

if weekly_sheets:
    print(f"\n'주간' 또는 'meeting'이 포함된 시트:")
    for name in weekly_sheets:
        print(f"  - {repr(name)}")
else:
    print("\n'주간' 또는 'meeting'이 포함된 시트를 찾을 수 없습니다.")

