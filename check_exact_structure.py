import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['경쟁사']

row = 3  # 롯데잠실

print("=== 롯데잠실의 Column 12 (L열) 주변 확인 ===")
print(f"Row 1 Column 12: {sheet.cell(row=1, column=12).value}")
print(f"Row 2 Column 12: {sheet.cell(row=2, column=12).value}")
print(f"Row 3 Column 12: {sheet.cell(row=3, column=12).value}")

print(f"\n=== Column 12부터 20까지 상세 확인 ===")
for col_idx in range(12, 21):
    header1 = sheet.cell(row=1, column=col_idx).value
    header2 = sheet.cell(row=2, column=col_idx).value
    value = sheet.cell(row=3, column=col_idx).value
    print(f"Column {col_idx}: Row1={header1}, Row2={header2}, Row3={value}")

# 387695 값이 정확히 어디에 있는지
print(f"\n=== 387695 값 찾기 (정확히) ===")
for col_idx in range(12, 30):
    value = sheet.cell(row=3, column=col_idx).value
    if isinstance(value, (int, float)):
        if 387690 <= value <= 387700:
            header1 = sheet.cell(row=1, column=col_idx).value
            header2 = sheet.cell(row=2, column=col_idx).value
            print(f"Column {col_idx}: {value} (Row 1: {header1}, Row 2: {header2})")


