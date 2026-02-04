import openpyxl
import json
from datetime import datetime

def read_excel_sheet_to_json(file_path, sheet_name, output_json_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            print(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
            print(f"사용 가능한 시트: {workbook.sheetnames}")
            return
        else:
            sheet = workbook[sheet_name]
        
        headers = [cell.value for cell in sheet[1]]
        data = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for header, value in zip(headers, row):
                if isinstance(value, datetime):
                    row_data[header] = value.strftime('%Y-%m-%d')
                else:
                    row_data[header] = value
            data.append(row_data)
        
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'headers': headers,
                'data': data,
                'total_rows': len(data)
            }, f, ensure_ascii=False, indent=2)
        
        print(f"총 {len(data)}개의 데이터를 {output_json_path}에 저장했습니다.")
        
        print("\n처음 5개 행 데이터:")
        for i, row in enumerate(data[:5]):
            print(f"행 {i+1}:")
            for key, value in row.items():
                print(f"  {key}: {value}")
            
    except FileNotFoundError:
        print(f"파일을 찾을 수 없습니다: {file_path}")
    except Exception as e:
        print(f"오류 발생: {e}")

excel_file = 'backdata.xlsx'
sheet_name = '아이템시즌별판매'
output_json_path = 'public/data/item_season_data.json'

read_excel_sheet_to_json(excel_file, sheet_name, output_json_path)





