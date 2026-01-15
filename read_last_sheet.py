import openpyxl
import json
from datetime import datetime

excel_file = 'backdata.xlsx'
workbook = openpyxl.load_workbook(excel_file, data_only=True)

print("모든 시트:")
for idx, name in enumerate(workbook.sheetnames, 1):
    print(f"  {idx}. {name}")

# 마지막 시트 확인
last_sheet_name = workbook.sheetnames[-1]
print(f"\n마지막 시트: {last_sheet_name}")

sheet = workbook[last_sheet_name]

# 헤더 확인
headers = [cell.value for cell in sheet[1]]
print(f"\n헤더 ({len(headers)}개):")
for idx, header in enumerate(headers, 1):
    print(f"  {idx}. {header}")

# 처음 3개 행 데이터 확인
print("\n처음 3개 행 데이터:")
for row_idx in range(2, min(5, sheet.max_row + 1)):
    print(f"\n행 {row_idx}:")
    for col_idx, header in enumerate(headers, 1):
        value = sheet.cell(row=row_idx, column=col_idx).value
        if value is not None:
            print(f"  {header}: {value}")

# '주간' 또는 '회의'가 헤더에 포함되어 있는지 확인
has_weekly_keyword = any('주간' in str(h) for h in headers) or any('회의' in str(h) for h in headers) if headers else False

if has_weekly_keyword or '주간' in last_sheet_name or '회의' in last_sheet_name:
    print("\n✓ 주간회의 관련 시트로 판단됩니다. JSON 변환을 진행합니다.")
    
    # 데이터 읽기
    data = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data = {}
        has_data = False
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d')
            elif value is None:
                value = None
            
            row_data[header] = value
            
            if value is not None and value != '':
                has_data = True
        
        if has_data:
            data.append(row_data)
    
    # JSON 저장
    output_json_path = 'weekly_meeting_data.json'
    result = {
        'headers': headers,
        'data': data,
        'total_rows': len(data)
    }
    
    with open(output_json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"\n✓ {len(data)}개 행을 {output_json_path}에 저장했습니다.")
else:
    print("\n⚠ 주간회의 시트가 아닌 것 같습니다. 시트 이름을 확인해주세요.")

