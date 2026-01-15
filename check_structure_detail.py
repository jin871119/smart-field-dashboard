import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['경쟁사']

# L열(12번째 컬럼)의 68행 확인
print("=== Row 68 데이터 확인 ===")
store_name = sheet.cell(row=68, column=11).value  # K열
mlb_value = sheet.cell(row=68, column=12).value   # L열
print(f"백화점 (K열): {store_name}")
print(f"MLB 월평균 (L열): {mlb_value}")

# 브랜드별로 점포별 데이터 확인
print("\n=== 브랜드별 점포 데이터 구조 확인 ===")
# Row 2에서 브랜드 찾기 (L열부터)
brands = []
for col_idx in range(12, 30):  # L열부터 확인
    brand_cell = sheet.cell(row=2, column=col_idx)
    if brand_cell.value:
        brands.append({
            'col': col_idx,
            'name': str(brand_cell.value).strip()
        })
        if len(brands) >= 5:  # 처음 5개만
            break

print(f"브랜드들: {[b['name'] for b in brands]}")

# 각 브랜드별로 68행의 값 확인
print("\n=== Row 68의 각 브랜드별 값 ===")
for brand in brands:
    value = sheet.cell(row=68, column=brand['col']).value
    print(f"{brand['name']} (Column {brand['col']}): {value}")

# 모든 점포의 MLB 데이터 확인 (순위 확인용)
print("\n=== MLB 브랜드의 모든 점포 데이터 (처음 10개) ===")
mlb_col = 12  # L열
for row_idx in range(3, 13):
    store = sheet.cell(row=row_idx, column=11).value
    value = sheet.cell(row=row_idx, column=mlb_col).value
    print(f"Row {row_idx}: {store} - {value}")



