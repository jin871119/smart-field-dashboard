import openpyxl
import json

def read_competitor_data(file_path, output_json_path):
    try:
        # data_only=True로 실제 계산된 값을 읽기
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['경쟁사']
        
        # Row 2에서 브랜드 목록 읽기 (L열 = 12번째 컬럼부터)
        brands = []
        brand_start_col = 12  # L열 (월평균 컬럼 시작)
        
        # 브랜드 헤더 찾기 (Row 2에서, L열부터)
        # 브랜드가 연속으로 있는 동안만 읽기 (빈 셀이 나오면 중단)
        for col_idx in range(brand_start_col, sheet.max_column + 1):
            brand_cell = sheet.cell(row=2, column=col_idx)
            if brand_cell.value and str(brand_cell.value).strip() and str(brand_cell.value).strip() != 'None':
                brand_name = str(brand_cell.value).strip()
                # Row 1의 헤더 확인 (월평균인지 확인)
                header_cell = sheet.cell(row=1, column=col_idx)
                if header_cell.value and '월평균' in str(header_cell.value):
                    brands.append({
                        'column': col_idx,
                        'name': brand_name
                    })
            elif len(brands) > 0:
                # 브랜드를 이미 찾았는데 빈 셀이 나오면 끝
                break
        
        print(f"발견된 브랜드: {[b['name'] for b in brands]}")
        
        # 데이터 읽기 (Row 3부터)
        stores_data = []
        
        for row_idx in range(3, sheet.max_row + 1):
            # 백화점 이름 (K열 = 11번째 컬럼)
            store_name = sheet.cell(row=row_idx, column=brand_start_col).value
            
            # 백화점 이름이 없으면 데이터 끝으로 간주
            if not store_name or str(store_name).strip() == '':
                break
            
            store_name = str(store_name).strip()
            
            # 각 브랜드별 월평균 데이터 수집
            brand_data = {}
            for brand in brands:
                cell = sheet.cell(row=row_idx, column=brand['column'])
                value = cell.value
                
                # 숫자 값 처리
                if value is None:
                    brand_data[brand['name']] = 0
                elif isinstance(value, (int, float)):
                    brand_data[brand['name']] = float(value)
                else:
                    # 문자열이면 숫자로 변환 시도
                    try:
                        brand_data[brand['name']] = float(value)
                    except:
                        brand_data[brand['name']] = 0
            
            stores_data.append({
                '백화점': store_name,
                '브랜드별_월평균': brand_data
            })
        
        # 결과 저장
        result = {
            'brands': [b['name'] for b in brands],
            'stores': stores_data,
            'total_stores': len(stores_data)
        }
        
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"\n총 {len(stores_data)}개 점포의 데이터를 {output_json_path}에 저장했습니다.")
        print(f"\n처음 3개 점포 데이터:")
        for store in stores_data[:3]:
            print(f"  {store['백화점']}: {store['브랜드별_월평균']}")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()

excel_file = 'backdata.xlsx'
output_json_path = 'competitor_data_v2.json'

read_competitor_data(excel_file, output_json_path)

