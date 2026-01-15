# -*- coding: utf-8 -*-
import openpyxl

excel_file = 'backdata.xlsx'

try:
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook['경쟁사']
    
    print("K열 (11번 컬럼) 데이터 확인:")
    print("\nRow 1-5:")
    for row_idx in range(1, 6):
        cell = sheet.cell(row=row_idx, column=11)
        print(f"  Row {row_idx}: {cell.value}")
    
    print("\nRow 4-15 (데이터 행):")
    for row_idx in range(4, 16):
        k_col = sheet.cell(row=row_idx, column=11).value
        # I열(9번)과 J열(10번)도 확인
        i_col = sheet.cell(row=row_idx, column=9).value
        j_col = sheet.cell(row=row_idx, column=10).value
        print(f"  Row {row_idx}: I={i_col}, J={j_col}, K={k_col}")
    
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()
