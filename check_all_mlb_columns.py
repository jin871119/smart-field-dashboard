import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['경쟁사']

# 롯데잠실 Row 3
row = 3

print("=== 롯데잠실 Row 3의 모든 MLB 관련 컬럼 확인 ===")

# Row 1에서 "월평균"이 있는 컬럼 찾기
monthly_avg_cols = []
for col_idx in range(1, sheet.max_column + 1):
    header_cell = sheet.cell(row=1, column=col_idx)
    if header_cell.value and '월평균' in str(header_cell.value):
        monthly_avg_cols.append(col_idx)

print(f"월평균 컬럼들: {monthly_avg_cols}")

# 각 월평균 컬럼의 Row 2 브랜드와 Row 3 값 확인
for start_col in monthly_avg_cols:
    print(f"\n=== Column {start_col}부터 확인 ===")
    for col_offset in range(0, 15):  # 15개 브랜드까지 확인
        col_idx = start_col + col_offset
        brand_cell = sheet.cell(row=2, column=col_idx)
        value_cell = sheet.cell(row=row, column=col_idx)
        
        if brand_cell.value:
            brand_name = str(brand_cell.value).strip()
            if brand_name == 'MLB':
                print(f"  Column {col_idx}: MLB = {value_cell.value}")
        
        if not brand_cell.value and col_offset > 0:
            break

# 특히 387695 값이 어디 있는지 찾기
print(f"\n=== 387695 값 찾기 ===")
for col_idx in range(1, sheet.max_column + 1):
    value = sheet.cell(row=row, column=col_idx).value
    if value and abs(float(value) - 387695) < 100:
        header1 = sheet.cell(row=1, column=col_idx).value
        header2 = sheet.cell(row=2, column=col_idx).value
        print(f"Column {col_idx}: {value} (Row 1: {header1}, Row 2: {header2})")


