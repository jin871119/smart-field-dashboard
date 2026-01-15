import openpyxl
import json

def read_group_data(file_path, output_json_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        if '단체' not in workbook.sheetnames:
            print("'단체' 시트를 찾을 수 없습니다.")
            return
        
        sheet = workbook['단체']
        
        # 매장별 소량단체판매액 데이터 수집
        stores_data = []
        
        for row_idx in range(2, sheet.max_row + 1):
            # 매장코드 (Column 1)
            store_code = sheet.cell(row=row_idx, column=1).value
            
            # 매장코드가 없으면 데이터 끝으로 간주
            if not store_code:
                break
            
            # 매장코드를 문자열로 변환
            store_code = str(store_code).strip()
            
            # 소량단체판매액 (T열 = Column 20)
            small_group_sales = sheet.cell(row=row_idx, column=20).value
            
            # 숫자 값 처리
            if small_group_sales is None:
                small_group_sales = 0
            elif isinstance(small_group_sales, (int, float)):
                small_group_sales = float(small_group_sales)
            else:
                try:
                    small_group_sales = float(small_group_sales)
                except:
                    small_group_sales = 0
            
            stores_data.append({
                '매장코드': store_code,
                '소량단체판매액': small_group_sales
            })
        
        # 결과 저장
        result = {
            'stores': stores_data,
            'total_stores': len(stores_data)
        }
        
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"총 {len(stores_data)}개 매장의 데이터를 {output_json_path}에 저장했습니다.")
        print(f"\n처음 5개 매장 데이터:")
        for store in stores_data[:5]:
            print(f"  매장코드 {store['매장코드']}: {store['소량단체판매액']:,.0f}원")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()

excel_file = 'backdata.xlsx'
output_json_path = 'group_sales_data.json'

read_group_data(excel_file, output_json_path)


