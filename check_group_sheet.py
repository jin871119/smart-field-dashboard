import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)

# 단체 시트 확인
if '단체' in wb.sheetnames:
    sheet = wb['단체']
    print("=== 단체 시트 발견 ===")
    
    # Row 1 헤더 확인
    print("\n=== Row 1 헤더 확인 ===")
    headers = []
    for col_idx in range(1, min(30, sheet.max_column + 1)):
        cell = sheet.cell(row=1, column=col_idx)
        if cell.value:
            headers.append((col_idx, cell.value))
            print(f"Column {col_idx}: {cell.value}")
    
    # T열 (20번째 컬럼) 확인
    print(f"\n=== T열 (Column 20) 확인 ===")
    t_col_header = sheet.cell(row=1, column=20).value
    print(f"Column 20 Header: {t_col_header}")
    
    # Row 2, 3 데이터 확인
    print(f"\n=== 처음 5개 행 데이터 확인 ===")
    for row_idx in range(2, 7):
        store_name = sheet.cell(row=row_idx, column=1).value if sheet.cell(row=row_idx, column=1).value else None
        t_col_value = sheet.cell(row=row_idx, column=20).value
        print(f"Row {row_idx}: 매장={store_name}, T열(소량단체판매액)={t_col_value}")
    
else:
    print("'단체' 시트를 찾을 수 없습니다.")
    print(f"사용 가능한 시트: {wb.sheetnames}")


