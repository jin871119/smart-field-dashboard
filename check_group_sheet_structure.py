import openpyxl

wb = openpyxl.load_workbook('backdata.xlsx', data_only=True)
sheet = wb['단체']

print("=== 단체 시트 구조 확인 ===")

# Row 1 헤더 확인
print("\n=== Row 1 헤더 (처음 30개 컬럼) ===")
for col_idx in range(1, min(31, sheet.max_column + 1)):
    cell = sheet.cell(row=1, column=col_idx)
    if cell.value:
        print(f"Column {col_idx}: {cell.value}")

# Row 2 확인
print("\n=== Row 2 확인 ===")
for col_idx in range(1, min(31, sheet.max_column + 1)):
    cell = sheet.cell(row=2, column=col_idx)
    if cell.value:
        print(f"Column {col_idx}: {cell.value}")

# 더현대서울 찾기
print("\n=== 더현대서울 데이터 찾기 ===")
for row_idx in range(2, min(100, sheet.max_row + 1)):
    store_name = sheet.cell(row=row_idx, column=2).value
    if store_name and '더현대서울' in str(store_name):
        print(f"\nRow {row_idx}: {store_name}")
        
        # T열 (20번째 컬럼) 값
        t_col_value = sheet.cell(row=row_idx, column=20).value
        print(f"  T열 (Column 20) 값: {t_col_value}")
        
        # 주변 컬럼 값들 확인
        print(f"\n  주변 컬럼 값들:")
        for col_idx in range(15, 26):
            header1 = sheet.cell(row=1, column=col_idx).value
            header2 = sheet.cell(row=2, column=col_idx).value
            value = sheet.cell(row=row_idx, column=col_idx).value
            if header1 or header2 or value:
                print(f"    Column {col_idx}: Row1={header1}, Row2={header2}, 값={value}")
        
        # 25년 월별 데이터가 있는지 확인 (202501, 202502 등의 컬럼)
        print(f"\n  25년 월별 데이터 컬럼 찾기:")
        for col_idx in range(1, sheet.max_column + 1):
            header1 = sheet.cell(row=1, column=col_idx).value
            if header1 and isinstance(header1, (int, float)):
                if 202501 <= header1 <= 202511:
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    print(f"    Column {col_idx} ({header1}): {value}")
            elif header1 and isinstance(header1, str):
                if '2025' in str(header1) or '25년' in str(header1):
                    value = sheet.cell(row=row_idx, column=col_idx).value
                    print(f"    Column {col_idx} ({header1}): {value}")


