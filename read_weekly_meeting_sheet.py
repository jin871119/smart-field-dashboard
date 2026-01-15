# -*- coding: utf-8 -*-
import openpyxl
import json
from datetime import datetime

def read_excel_sheet_to_json(file_path, sheet_name, output_json_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            print(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
            print(f"사용 가능한 시트: {workbook.sheetnames}")
            return False
        else:
            sheet = workbook[sheet_name]
        
        # 헤더 읽기 (1행)
        headers = [cell.value for cell in sheet[1]]
        print(f"헤더 ({len(headers)}개): {headers}")
        
        # 데이터 읽기 (2행부터)
        data = []
        
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # 날짜 타입 처리
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                elif value is None:
                    value = None
                
                row_data[header] = value
                
                # 값이 있으면 데이터가 있다고 표시
                if value is not None and value != '':
                    has_data = True
            
            # 데이터가 있는 행만 추가
            if has_data:
                data.append(row_data)
        
        # 결과 저장
        result = {
            'headers': headers,
            'data': data,
            'total_rows': len(data)
        }
        
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"\n총 {len(data)}개 행의 데이터를 {output_json_path}에 저장했습니다.")
        
        print("\n처음 3개 행 데이터:")
        for i, row in enumerate(data[:3]):
            print(f"\n행 {i+1}:")
            for key, value in list(row.items())[:10]:  # 처음 10개 컬럼만
                print(f"  {key}: {value}")
        
        return True
        
    except FileNotFoundError:
        print(f"파일을 찾을 수 없습니다: {file_path}")
        return False
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False

# 메인 실행
excel_file = 'backdata.xlsx'
sheet_name = '주간회의'
output_json_path = 'weekly_meeting_data.json'

print(f"'{sheet_name}' 시트를 읽는 중...")
success = read_excel_sheet_to_json(excel_file, sheet_name, output_json_path)

if not success:
    print("\n시트를 찾을 수 없습니다. 엑셀 파일이 최신 버전인지 확인해주세요.")
    print("또는 시트 이름이 다른 경우 정확한 시트 이름을 알려주세요.")

