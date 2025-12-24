import openpyxl
import json

def read_competitor_data(file_path, output_json_path):
    try:
        # data_only=True로 실제 계산된 값을 읽기
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['경쟁사']
        
        # Row 1에서 "월평균(1~11월)" 헤더가 있는 첫 번째 컬럼 찾기
        monthly_avg_start_col = None
        for col_idx in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=1, column=col_idx)
            if header_cell.value and '월평균' in str(header_cell.value):
                monthly_avg_start_col = col_idx
                break
        
        if not monthly_avg_start_col:
            print("월평균 컬럼을 찾을 수 없습니다.")
            return
        
        print(f"월평균 컬럼 시작: Column {monthly_avg_start_col} (L열)")
        
        # Row 2에서 브랜드 목록 읽기 (월평균 시작 컬럼부터)
        brands = []
        seen_brands = set()  # 이미 본 브랜드 추적
        
        for col_idx in range(monthly_avg_start_col, sheet.max_column + 1):
            brand_cell = sheet.cell(row=2, column=col_idx)
            if brand_cell.value and str(brand_cell.value).strip():
                brand_name = str(brand_cell.value).strip()
                
                # 같은 브랜드가 이미 있으면 건너뛰기 (첫 번째 것만 사용)
                if brand_name in seen_brands:
                    continue
                
                # 실제 데이터가 있는지 확인 (Row 3에 숫자 값이 있는지)
                test_cell = sheet.cell(row=3, column=col_idx)
                if test_cell.value is not None and isinstance(test_cell.value, (int, float)):
                    brands.append({
                        'column': col_idx,
                        'name': brand_name
                    })
                    seen_brands.add(brand_name)
            elif len(brands) > 0:
                # 브랜드를 이미 찾았는데 빈 셀이 나오면 끝
                # 하지만 더 있을 수 있으니 일단 계속 확인
                continue
        
        print(f"발견된 브랜드 ({len(brands)}개): {[b['name'] for b in brands[:10]]}...")
        
        # 데이터 읽기 (Row 3부터)
        stores_data = []
        
        for row_idx in range(3, sheet.max_row + 1):
            # 백화점 이름 (K열 = 11번째 컬럼)
            store_name_cell = sheet.cell(row=row_idx, column=11)
            store_name = store_name_cell.value
            
            # 백화점 이름이 없거나 숫자인 경우 데이터 끝으로 간주
            if not store_name:
                break
            if isinstance(store_name, (int, float)):
                break
            
            store_name = str(store_name).strip()
            if store_name == '':
                break
            
            # 각 브랜드별 월평균 데이터 수집
            brand_data = {}
            for brand in brands:
                cell = sheet.cell(row=row_idx, column=brand['column'])
                value = cell.value
                
                # 숫자 값 처리
                if value is None:
                    brand_data[brand['name']] = 0
                elif isinstance(value, (int, float)):
                    brand_data[brand['name']] = float(value)
                else:
                    # 문자열이면 숫자로 변환 시도
                    try:
                        brand_data[brand['name']] = float(value)
                    except:
                        brand_data[brand['name']] = 0
            
            stores_data.append({
                '백화점': store_name,
                '브랜드별_월평균': brand_data
            })
        
        # 브랜드별 순위 계산
        brand_rankings = {}
        for brand in brands:
            # 해당 브랜드의 점포별 데이터 수집
            store_brand_data = []
            for store_data in stores_data:
                brand_value = store_data['브랜드별_월평균'].get(brand['name'], 0)
                if brand_value > 0:  # 0보다 큰 값만 포함
                    store_brand_data.append({
                        '백화점': store_data['백화점'],
                        '월평균': brand_value
                    })
            
            # 월평균 기준 내림차순 정렬
            store_brand_data.sort(key=lambda x: x['월평균'], reverse=True)
            
            # 순위 추가
            for idx, item in enumerate(store_brand_data):
                item['순위'] = idx + 1
            
            brand_rankings[brand['name']] = store_brand_data
        
        # 결과 저장
        result = {
            'brands': [b['name'] for b in brands],
            'stores': stores_data,
            'total_stores': len(stores_data),
            'brand_rankings': brand_rankings  # 브랜드별 순위 추가
        }
        
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        print(f"\n총 {len(stores_data)}개 점포의 데이터를 {output_json_path}에 저장했습니다.")
        print(f"\n처음 3개 점포 데이터:")
        for store in stores_data[:3]:
            print(f"  {store['백화점']}:")
            for brand, value in list(store['브랜드별_월평균'].items())[:5]:
                print(f"    {brand}: {value:,.0f}")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()

excel_file = 'backdata.xlsx'
output_json_path = 'competitor_data_v2.json'

read_competitor_data(excel_file, output_json_path)

