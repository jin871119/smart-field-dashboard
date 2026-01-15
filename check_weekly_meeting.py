import openpyxl
import json

excel_file = 'backdata.xlsx'
sheet_name = '주간회의'

try:
    workbook = openpyxl.load_workbook(excel_file)
    
    print(f"사용 가능한 시트: {workbook.sheetnames}")
    
    if sheet_name not in workbook.sheetnames:
        print(f"\n'{sheet_name}' 시트를 찾을 수 없습니다.")
    else:
        sheet = workbook[sheet_name]
        
        print(f"\n'{sheet_name}' 시트 구조:")
        print(f"총 행 수: {sheet.max_row}")
        print(f"총 열 수: {sheet.max_column}")
        
        # 헤더 확인 (1행)
        headers = [cell.value for cell in sheet[1]]
        print(f"\n헤더 ({len(headers)}개):")
        for idx, header in enumerate(headers, 1):
            print(f"  {idx}. {header}")
        
        # 처음 5개 행 데이터 확인
        print("\n처음 5개 행 데이터:")
        for row_idx in range(2, min(7, sheet.max_row + 1)):
            print(f"\n행 {row_idx}:")
            for col_idx, header in enumerate(headers, 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                print(f"  {header}: {value}")

except FileNotFoundError:
    print(f"파일을 찾을 수 없습니다: {excel_file}")
except Exception as e:
    print(f"오류 발생: {e}")
    import traceback
    traceback.print_exc()

